"""Phase 6.11a: tests for the /report multi-month summary."""

from __future__ import annotations

from collections.abc import AsyncIterator
from datetime import date
from decimal import Decimal
from zoneinfo import ZoneInfo

import pytest
import pytest_asyncio
from openpyxl import Workbook
from sqlalchemy import BigInteger
from sqlalchemy.ext.asyncio import (
    AsyncSession,
    async_sessionmaker,
    create_async_engine,
)
from sqlalchemy.ext.compiler import compiles
from src.bot.handlers.report import parse_months_arg
from src.core.models import Advance, Crew, DayEntry, SalaryPayment, User
from src.services import accounting as accounting_module
from src.services.advances import SalaryBreakdown, record_advance
from src.services.reports.archive import archive_filename, build_report_archive
from src.services.reports.pdf import build_report_pdf, pdf_filename
from src.services.reports.png import build_report_png, png_filename
from src.services.reports.service import get_report_data
from src.services.reports.text import format_report_text
from src.services.reports.xlsx import build_report_xlsx, xlsx_filename
from src.services.salary_payments import record_payment


@compiles(BigInteger, "sqlite")
def _bigint_as_integer_on_sqlite(*_args: object, **_kw: object) -> str:
    return "INTEGER"


_TABLES_FOR_TESTS = [
    Crew.__table__,
    User.__table__,
    DayEntry.__table__,
    Advance.__table__,
    SalaryPayment.__table__,
]

_TZ = ZoneInfo("Europe/Warsaw")


@pytest_asyncio.fixture
async def session() -> AsyncIterator[AsyncSession]:
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", future=True)
    async with engine.begin() as conn:
        for table in _TABLES_FOR_TESTS:
            await conn.run_sync(table.create)  # type: ignore[attr-defined]
    factory = async_sessionmaker(engine, expire_on_commit=False)
    async with factory() as s:
        yield s
    await engine.dispose()


async def _seed_user(
    session: AsyncSession, *, rate: Decimal | None = Decimal("50.00"),
) -> User:
    user = User(tg_id=100, name="Worker", hourly_rate=rate, currency="PLN")
    session.add(user)
    await session.flush()
    return user


def _stub_compute_salary(
    monkeypatch: pytest.MonkeyPatch,
    *,
    by_month: dict[tuple[int, int], tuple[Decimal, Decimal | None]],
) -> None:
    """Replace compute_salary with a per-(year, month) lookup table."""
    async def fake(
        _session: AsyncSession, *, user: User, year: int, month: int,
        tz: ZoneInfo,  # noqa: ARG001
    ) -> SalaryBreakdown:
        hours, earnings = by_month.get((year, month), (Decimal(0), None))
        return SalaryBreakdown(
            user_id=user.id,
            year=year, month=month,
            day_entries_hours=hours,
            day_entries_earnings=earnings,
            shifts_hours=Decimal(0),
            shifts_earnings=None,
            advances_total=Decimal(0),
            net_payable=earnings,
        )
    monkeypatch.setattr(accounting_module, "compute_salary", fake)


# --- parse_months_arg --------------------------------------------------


def test_parse_months_default_when_none() -> None:
    assert parse_months_arg(None) == 6


def test_parse_months_default_when_empty_string() -> None:
    assert parse_months_arg("") == 6
    assert parse_months_arg("   ") == 6


def test_parse_months_valid_value() -> None:
    assert parse_months_arg("12") == 12
    assert parse_months_arg(" 1 ") == 1
    assert parse_months_arg("24") == 24


def test_parse_months_out_of_range_rejected() -> None:
    assert parse_months_arg("0") is None
    assert parse_months_arg("25") is None
    assert parse_months_arg("-3") is None


def test_parse_months_non_numeric_rejected() -> None:
    assert parse_months_arg("abc") is None
    assert parse_months_arg("3.5") is None


# --- get_report_data ---------------------------------------------------


async def test_report_data_walks_back_n_months(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
        (2026, 4): (Decimal("160"), Decimal("8000")),
        (2026, 3): (Decimal("140"), Decimal("7000")),
    })
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 12), months=3,
    )
    assert data.months == 3
    assert [(lg.year, lg.month) for lg in data.ledgers] == [
        (2026, 5), (2026, 4), (2026, 3),
    ]
    assert data.total_hours == Decimal("460.00")
    assert data.total_earnings == Decimal("23000.00")


async def test_report_data_total_owed_sums_positive_remainings(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
        (2026, 4): (Decimal("160"), Decimal("8000")),
    })
    # April fully paid; May still owed.
    await record_payment(
        session, user_id=user.id, paid_on=date(2026, 5, 5),
        period_year=2026, period_month=4, amount=Decimal("8000"),
        recorded_by_id=user.id,
    )
    await session.commit()
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=2,
    )
    assert data.total_received == Decimal("8000.00")
    assert data.total_owed == Decimal("8000.00")  # only May
    assert data.total_overpaid == Decimal("0.00")


async def test_report_data_total_overpaid_when_received_exceeds_earnings(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
    })
    await record_advance(
        session, user_id=user.id, amount=Decimal("9000"),
        recorded_by_id=user.id, day=date(2026, 5, 1),
        period_year=2026, period_month=5,
    )
    await session.commit()
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 31), months=1,
    )
    assert data.total_owed == Decimal("0.00")
    assert data.total_overpaid == Decimal("1000.00")


async def test_report_data_unpriced_months_dont_break_totals(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session, rate=None)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("100"), None),
        (2026, 4): (Decimal("80"), None),
    })
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 12), months=2,
    )
    assert data.total_hours == Decimal("180.00")
    assert data.total_earnings is None
    assert data.total_owed == Decimal("0.00")


# --- format_report_text ------------------------------------------------


async def test_format_report_text_contains_period_labels_and_totals(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
        (2026, 4): (Decimal("160"), Decimal("8000")),
    })
    await record_payment(
        session, user_id=user.id, paid_on=date(2026, 5, 5),
        period_year=2026, period_month=4, amount=Decimal("8000"),
        recorded_by_id=user.id,
    )
    await session.commit()
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=2,
    )
    text = format_report_text(data, user)
    assert "Май 2026" in text
    assert "Апрель 2026" in text
    assert "PLN" in text
    # Totals row mentions debt (долг) including the May 8000 PLN.
    assert "долг 8000.00" in text


async def test_format_report_text_renders_unpriced_row(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session, rate=None)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("100"), None),
    })
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=1,
    )
    text = format_report_text(data, user)
    assert "без ставки" in text


# --- build_report_xlsx -------------------------------------------------


def _load_workbook(buf: bytes) -> Workbook:
    from io import BytesIO

    from openpyxl import load_workbook

    return load_workbook(BytesIO(buf), data_only=True)


def test_xlsx_filename_includes_window() -> None:
    assert xlsx_filename(6) == "report_6m.xlsx"
    assert xlsx_filename(12) == "report_12m.xlsx"


async def test_build_report_xlsx_has_header_period_and_totals(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
        (2026, 4): (Decimal("160"), Decimal("8000")),
    })
    await record_payment(
        session, user_id=user.id, paid_on=date(2026, 5, 5),
        period_year=2026, period_month=4, amount=Decimal("8000"),
        recorded_by_id=user.id,
    )
    await session.commit()
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=2,
    )
    buf = build_report_xlsx(data, user)
    wb = _load_workbook(buf.getvalue())
    ws = wb["Отчёт"]
    # Title on row 1.
    assert "2 мес" in str(ws.cell(row=1, column=1).value)
    # Column headers on row 3 include the currency code.
    headers = [ws.cell(row=3, column=c).value for c in range(1, 7)]
    assert headers[0] == "Период"
    assert "PLN" in str(headers[2])  # Начислено (PLN)
    # First data row is the newest month (May 2026).
    assert ws.cell(row=4, column=1).value == "Май 2026"
    assert ws.cell(row=4, column=2).value == 160.0
    assert ws.cell(row=4, column=3).value == 8000.0
    # Second data row is April 2026 (paid in full → settled).
    assert ws.cell(row=5, column=1).value == "Апрель 2026"
    assert "закрыт" in str(ws.cell(row=5, column=6).value)
    # Totals row sits one blank line below the data block (row 7).
    assert ws.cell(row=7, column=1).value == "Итого"
    assert ws.cell(row=7, column=2).value == 320.0
    assert ws.cell(row=7, column=4).value == 8000.0
    assert ws.cell(row=7, column=5).value == 8000.0  # outstanding debt


async def test_build_report_xlsx_renders_unpriced_dash(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session, rate=None)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("100"), None),
    })
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=1,
    )
    buf = build_report_xlsx(data, user)
    wb = _load_workbook(buf.getvalue())
    ws = wb["Отчёт"]
    # Earnings + remaining cells fall back to em dash for unpriced rows.
    assert ws.cell(row=4, column=3).value == "—"
    assert ws.cell(row=4, column=5).value == "—"
    assert "без ставки" in str(ws.cell(row=4, column=6).value)


async def test_build_report_xlsx_shows_overpaid_row(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
    })
    await record_payment(
        session, user_id=user.id, paid_on=date(2026, 5, 5),
        period_year=2026, period_month=5, amount=Decimal("9000"),
        recorded_by_id=user.id,
    )
    await session.commit()
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=1,
    )
    buf = build_report_xlsx(data, user)
    wb = _load_workbook(buf.getvalue())
    ws = wb["Отчёт"]
    # Totals row 6 (one data row), overpaid row directly below.
    assert ws.cell(row=6, column=1).value == "Итого"
    assert ws.cell(row=7, column=1).value == "Переплата"
    assert ws.cell(row=7, column=5).value == 1000.0


# --- build_report_pdf --------------------------------------------------


def test_pdf_filename_includes_window() -> None:
    assert pdf_filename(6) == "report_6m.pdf"
    assert pdf_filename(12) == "report_12m.pdf"


async def test_build_report_pdf_returns_valid_pdf_bytes(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
        (2026, 4): (Decimal("160"), Decimal("8000")),
    })
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=2,
    )
    buf = build_report_pdf(data, user)
    blob = buf.getvalue()
    # PDF magic header + plausible size for a one-page document.
    assert blob.startswith(b"%PDF-")
    assert blob.rstrip().endswith(b"%%EOF")
    assert len(blob) > 2000


async def test_build_report_pdf_handles_overpaid(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
    })
    await record_payment(
        session, user_id=user.id, paid_on=date(2026, 5, 5),
        period_year=2026, period_month=5, amount=Decimal("9000"),
        recorded_by_id=user.id,
    )
    await session.commit()
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=1,
    )
    buf = build_report_pdf(data, user)
    assert buf.getvalue().startswith(b"%PDF-")


# --- build_report_png --------------------------------------------------


_PNG_MAGIC = b"\x89PNG\r\n\x1a\n"


def test_png_filename_includes_window() -> None:
    assert png_filename(6) == "report_6m.png"
    assert png_filename(24) == "report_24m.png"


async def test_build_report_png_returns_valid_png_bytes(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
        (2026, 4): (Decimal("160"), Decimal("8000")),
    })
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=2,
    )
    buf = build_report_png(data, user)
    blob = buf.getvalue()
    assert blob.startswith(_PNG_MAGIC)
    assert len(blob) > 2000


# --- build_report_archive ----------------------------------------------


def test_archive_filename_includes_window() -> None:
    assert archive_filename(6) == "report_6m.zip"
    assert archive_filename(24) == "report_24m.zip"


async def test_build_report_archive_contains_all_three_formats(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    import zipfile
    from io import BytesIO

    user = await _seed_user(session)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("160"), Decimal("8000")),
        (2026, 4): (Decimal("160"), Decimal("8000")),
    })
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=2,
    )
    buf = build_report_archive(data, user, months=2)
    with zipfile.ZipFile(BytesIO(buf.getvalue())) as zf:
        names = set(zf.namelist())
        assert names == {"report_2m.xlsx", "report_2m.pdf", "report_2m.png"}
        assert zf.read("report_2m.pdf").startswith(b"%PDF-")
        assert zf.read("report_2m.png").startswith(_PNG_MAGIC)
        # XLSX is a ZIP itself; sniff via openpyxl rather than magic bytes.
        wb = _load_workbook(zf.read("report_2m.xlsx"))
        assert "Отчёт" in wb.sheetnames


async def test_build_report_png_handles_unpriced_months(
    session: AsyncSession, monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = await _seed_user(session, rate=None)
    _stub_compute_salary(monkeypatch, by_month={
        (2026, 5): (Decimal("100"), None),
        (2026, 4): (Decimal("80"), None),
    })
    data = await get_report_data(
        session, user=user, tz=_TZ, today=date(2026, 5, 20), months=2,
    )
    buf = build_report_png(data, user)
    assert buf.getvalue().startswith(_PNG_MAGIC)
