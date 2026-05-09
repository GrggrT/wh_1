"""Tests for XLSX exporter."""

from datetime import datetime
from decimal import Decimal
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from src.exporters.xlsx import build_shift_rows, export_crew_xlsx, export_xlsx

from tests.conftest import FakeShift, FakeSite, FakeUser


class TestBuildShiftRows:
    def test_normal_shift(
        self,
        sample_shifts: list[FakeShift],
        sample_site: FakeSite,
        sample_user: FakeUser,
        tz_warsaw: ZoneInfo,
    ) -> None:
        sites: dict[int, Any] = {1: sample_site}
        rows = build_shift_rows(
            [sample_shifts[0]], sites, sample_user, tz_warsaw,  # type: ignore[arg-type]
        )
        assert len(rows) == 1
        assert rows[0].hours == Decimal("8.00")
        assert rows[0].rate == Decimal("60.00")
        assert rows[0].amount == Decimal("480.00")

    def test_cross_midnight_split(
        self,
        sample_shifts: list[FakeShift],
        sample_site: FakeSite,
        sample_user: FakeUser,
        tz_warsaw: ZoneInfo,
    ) -> None:
        sites: dict[int, Any] = {1: sample_site}
        rows = build_shift_rows(
            [sample_shifts[1]], sites, sample_user, tz_warsaw,  # type: ignore[arg-type]
        )
        assert len(rows) == 2
        total_hours = sum(r.hours for r in rows)
        assert total_hours == Decimal("1.00")

    def test_user_rate_fallback(
        self,
        sample_shifts: list[FakeShift],
        sample_user: FakeUser,
        tz_warsaw: ZoneInfo,
    ) -> None:
        """When site has no rate, user rate is used."""
        site_no_rate = FakeSite(name="No Rate Site", hourly_rate=None)
        sites: dict[int, Any] = {1: site_no_rate}
        rows = build_shift_rows(
            [sample_shifts[0]], sites, sample_user, tz_warsaw,  # type: ignore[arg-type]
        )
        assert rows[0].rate == Decimal("50.00")


class TestExportXlsx:
    def test_produces_valid_workbook(
        self,
        sample_shifts: list[FakeShift],
        sample_site: FakeSite,
        sample_user: FakeUser,
        tz_warsaw: ZoneInfo,
    ) -> None:
        sites: dict[int, Any] = {1: sample_site}
        buffer = export_xlsx(
            sample_shifts, sites, sample_user, tz_warsaw, "2026-05",  # type: ignore[arg-type]
        )

        wb = load_workbook(buffer)
        assert "Shifts" in wb.sheetnames
        assert "Summary" in wb.sheetnames

        ws = wb["Shifts"]
        # Header + 3 rows (8hr shift + 2 rows from midnight split)
        assert ws.max_row == 4

        # Check hours sum
        total_hours = Decimal(0)
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=5).value
            if val is not None:
                total_hours += Decimal(str(val))

        # 8h + 1h = 9h
        assert total_hours == Decimal("9.00")

    def test_empty_shifts(
        self, sample_user: FakeUser, tz_warsaw: ZoneInfo,
    ) -> None:
        buffer = export_xlsx(
            [], {}, sample_user, tz_warsaw, "2026-05",  # type: ignore[arg-type]
        )
        wb = load_workbook(buffer)
        ws = wb["Shifts"]
        assert ws.max_row == 1  # Only header


class TestExportCrewXlsx:
    def test_summary_and_per_member_sheets(self, tz_warsaw: ZoneInfo) -> None:
        crew = type("FakeCrew", (), {"id": 1, "name": "Бригада А"})()

        m1 = FakeUser(id=10, name="Иван")
        m2 = FakeUser(id=20, name="Петр")
        m1.hourly_rate = Decimal("50.00")  # type: ignore[misc]
        m2.hourly_rate = Decimal("60.00")  # type: ignore[misc]

        s1 = FakeShift(
            id=1, user_id=10, site_id=None,
            start_at=datetime(2026, 5, 8, 8, 0, tzinfo=tz_warsaw),
            end_at=datetime(2026, 5, 8, 16, 0, tzinfo=tz_warsaw),
        )
        s2 = FakeShift(
            id=2, user_id=20, site_id=None,
            start_at=datetime(2026, 5, 8, 9, 0, tzinfo=tz_warsaw),
            end_at=datetime(2026, 5, 8, 13, 0, tzinfo=tz_warsaw),
        )

        buffer = export_crew_xlsx(
            crew,  # type: ignore[arg-type]
            [m1, m2],  # type: ignore[list-item]
            {10: [s1], 20: [s2]},  # type: ignore[dict-item]
            {},
            tz_warsaw,
            "2026-05",
        )
        wb = load_workbook(buffer)
        assert "Crew summary" in wb.sheetnames
        assert "Иван" in wb.sheetnames
        assert "Петр" in wb.sheetnames

        sumsheet = wb["Crew summary"]
        member_rows = [
            sumsheet.cell(row=r, column=1).value
            for r in range(4, sumsheet.max_row + 1)
        ]
        assert "Иван" in member_rows
        assert "Петр" in member_rows
        assert "Grand total" in member_rows
