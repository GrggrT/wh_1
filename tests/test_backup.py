"""Phase 7.1: tests for the full-data XLSX backup builder."""

from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from src.core.models import Advance, DayEntry, SalaryPayment, User
from src.services.reports.backup import (
    backup_filename,
    build_backup_xlsx,
)


def _user() -> User:
    return User(
        id=42,
        tg_id=12345,
        name="Иван",
        locale="ru",
        hourly_rate=Decimal("30.00"),
        currency="PLN",
        role="worker",
    )


def _day(user_id: int, d: date, hours: str) -> DayEntry:
    return DayEntry(
        id=1, user_id=user_id, day=d, hours=Decimal(hours), note=None,
        created_at=datetime(2026, 5, 1, 12, 0, tzinfo=UTC),
        updated_at=datetime(2026, 5, 1, 12, 0, tzinfo=UTC),
    )


def _advance(user_id: int, d: date, amount: str) -> Advance:
    return Advance(
        id=1, user_id=user_id, day=d,
        period_year=d.year, period_month=d.month,
        amount=Decimal(amount), note="zaliczka", recorded_by_id=user_id,
        created_at=datetime(2026, 5, 1, 12, 0, tzinfo=UTC),
    )


def _payment(user_id: int, d: date, amount: str) -> SalaryPayment:
    return SalaryPayment(
        id=1, user_id=user_id, paid_on=d,
        period_year=d.year, period_month=d.month,
        amount=Decimal(amount), note=None, recorded_by_id=user_id,
        created_at=datetime(2026, 5, 1, 12, 0, tzinfo=UTC),
    )


def _load(buf) -> Workbook:  # noqa: ANN001
    buf.seek(0)
    return load_workbook(buf)


def test_backup_filename_includes_user_id_and_date() -> None:
    name = backup_filename(_user(), date(2026, 5, 12))
    assert name == "wh1_backup_42_2026-05-12.xlsx"


def test_backup_xlsx_has_four_sheets() -> None:
    buf = build_backup_xlsx(
        _user(), [], [], [], today=date(2026, 5, 12),
    )
    wb = _load(buf)
    assert wb.sheetnames == ["Профиль", "Дни", "Авансы", "Выплаты"]


def test_backup_profile_sheet_lists_key_fields() -> None:
    buf = build_backup_xlsx(
        _user(), [], [], [], today=date(2026, 5, 12),
    )
    ws = _load(buf)["Профиль"]
    labels = [ws.cell(row=r, column=1).value for r in range(1, 8)]
    assert "user_id" in labels
    assert "tg_id" in labels
    assert "currency" in labels
    assert "hourly_rate" in labels
    assert "backup_date" in labels


def test_backup_data_rows_render() -> None:
    user = _user()
    days = [_day(user.id, date(2026, 5, 1), "8")]
    advs = [_advance(user.id, date(2026, 5, 5), "100.00")]
    pays = [_payment(user.id, date(2026, 5, 10), "500.00")]
    buf = build_backup_xlsx(user, days, advs, pays, today=date(2026, 5, 12))
    wb = _load(buf)

    ws_days = wb["Дни"]
    assert ws_days.cell(row=2, column=3).value == 8.0  # hours

    ws_adv = wb["Авансы"]
    assert ws_adv.cell(row=2, column=5).value == 100.0  # amount

    ws_pay = wb["Выплаты"]
    assert ws_pay.cell(row=2, column=5).value == 500.0  # amount


def test_backup_empty_sheets_have_headers_only() -> None:
    buf = build_backup_xlsx(
        _user(), [], [], [], today=date(2026, 5, 12),
    )
    wb = _load(buf)
    for sheet in ("Дни", "Авансы", "Выплаты"):
        ws = wb[sheet]
        # Header row is row 1, data starts at row 2; only header populated.
        assert ws.cell(row=1, column=1).value is not None
        assert ws.cell(row=2, column=1).value is None
