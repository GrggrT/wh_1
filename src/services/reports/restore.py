"""Phase 7.1b: restore a user's data from a /backup XLSX.

Mirror of :mod:`src.services.reports.backup`. Parses the 4-sheet workbook
produced by ``/backup`` and inserts the rows back into the accounting
tables for the calling user.

Row-level conflict policy (intentional — restore should be idempotent
and never overwrite existing data):

* ``day_entries`` — UNIQUE on ``(user_id, day)``. Existing rows are kept
  as-is; the incoming hours value is skipped.
* ``advances`` / ``salary_payments`` — no natural unique key in the
  schema. An incoming row is treated as a duplicate when an existing row
  matches on ``(user_id, day/paid_on, period_year, period_month, amount,
  note)``. Anything that doesn't match is inserted as a new row.

Original primary keys from the source workbook are ignored — the new
deployment assigns its own IDs.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.core.models import Advance, DayEntry, SalaryPayment, User


@dataclass
class DayEntryRow:
    day: date
    hours: Decimal
    note: str | None


@dataclass
class AdvanceRow:
    day: date
    period_year: int
    period_month: int
    amount: Decimal
    note: str | None


@dataclass
class PaymentRow:
    paid_on: date
    period_year: int
    period_month: int
    amount: Decimal
    note: str | None


@dataclass
class RestorePlan:
    days: list[DayEntryRow]
    advances: list[AdvanceRow]
    payments: list[PaymentRow]


@dataclass
class RestoreResult:
    days_inserted: int
    days_skipped: int
    advances_inserted: int
    advances_skipped: int
    payments_inserted: int
    payments_skipped: int


class BackupParseError(ValueError):
    """Raised when the uploaded workbook doesn't match the backup layout."""


_REQUIRED_SHEETS = ("Дни", "Авансы", "Выплаты")


def _to_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return date.fromisoformat(value.strip())
    raise BackupParseError(f"expected a date, got {value!r}")


def _to_decimal(value: object) -> Decimal:
    if value is None or value == "":
        raise BackupParseError("missing required amount")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise BackupParseError(f"invalid decimal: {value!r}") from exc


def _to_int(value: object) -> int:
    if isinstance(value, bool):  # bool is an int subclass
        raise BackupParseError(f"expected int, got bool {value!r}")
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip():
        return int(value.strip())
    raise BackupParseError(f"expected an integer, got {value!r}")


def _opt_note(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def parse_backup_xlsx(buf: BytesIO) -> RestorePlan:
    """Parse the 4-sheet backup workbook into a :class:`RestorePlan`."""
    wb = load_workbook(buf, data_only=True, read_only=True)
    missing = [name for name in _REQUIRED_SHEETS if name not in wb.sheetnames]
    if missing:
        raise BackupParseError(f"missing sheets: {missing}")

    days: list[DayEntryRow] = []
    for row in wb["Дни"].iter_rows(min_row=2, values_only=True):
        if row is None or row[1] is None:
            continue
        days.append(DayEntryRow(
            day=_to_date(row[1]),
            hours=_to_decimal(row[2]),
            note=_opt_note(row[4] if len(row) > 4 else None),
        ))

    advances: list[AdvanceRow] = []
    for row in wb["Авансы"].iter_rows(min_row=2, values_only=True):
        if row is None or row[1] is None:
            continue
        advances.append(AdvanceRow(
            day=_to_date(row[1]),
            period_year=_to_int(row[2]),
            period_month=_to_int(row[3]),
            amount=_to_decimal(row[4]),
            note=_opt_note(row[5] if len(row) > 5 else None),
        ))

    payments: list[PaymentRow] = []
    for row in wb["Выплаты"].iter_rows(min_row=2, values_only=True):
        if row is None or row[1] is None:
            continue
        payments.append(PaymentRow(
            paid_on=_to_date(row[1]),
            period_year=_to_int(row[2]),
            period_month=_to_int(row[3]),
            amount=_to_decimal(row[4]),
            note=_opt_note(row[5] if len(row) > 5 else None),
        ))

    return RestorePlan(days=days, advances=advances, payments=payments)


async def apply_restore(
    session: AsyncSession, *, user: User, plan: RestorePlan,
) -> RestoreResult:
    """Insert rows from ``plan`` for ``user``, skipping duplicates."""
    # Existing day keys for the unique-constraint guard.
    existing_days: set[date] = set(
        (await session.execute(
            select(DayEntry.day).where(DayEntry.user_id == user.id),
        )).scalars().all(),
    )
    days_inserted = days_skipped = 0
    for day_row in plan.days:
        if day_row.day in existing_days:
            days_skipped += 1
            continue
        session.add(DayEntry(
            user_id=user.id, day=day_row.day,
            hours=day_row.hours, note=day_row.note,
        ))
        existing_days.add(day_row.day)
        days_inserted += 1

    existing_advances: set[tuple[date, int, int, Decimal, str | None]] = {
        (a.day, a.period_year, a.period_month, a.amount, a.note)
        for a in (await session.execute(
            select(Advance).where(Advance.user_id == user.id),
        )).scalars().all()
    }
    advances_inserted = advances_skipped = 0
    for adv_row in plan.advances:
        adv_key = (
            adv_row.day, adv_row.period_year, adv_row.period_month,
            adv_row.amount, adv_row.note,
        )
        if adv_key in existing_advances:
            advances_skipped += 1
            continue
        session.add(Advance(
            user_id=user.id, day=adv_row.day,
            period_year=adv_row.period_year, period_month=adv_row.period_month,
            amount=adv_row.amount, note=adv_row.note,
            recorded_by_id=user.id,
        ))
        existing_advances.add(adv_key)
        advances_inserted += 1

    existing_payments: set[tuple[date, int, int, Decimal, str | None]] = {
        (p.paid_on, p.period_year, p.period_month, p.amount, p.note)
        for p in (await session.execute(
            select(SalaryPayment).where(SalaryPayment.user_id == user.id),
        )).scalars().all()
    }
    payments_inserted = payments_skipped = 0
    for pay_row in plan.payments:
        pay_key = (
            pay_row.paid_on, pay_row.period_year, pay_row.period_month,
            pay_row.amount, pay_row.note,
        )
        if pay_key in existing_payments:
            payments_skipped += 1
            continue
        session.add(SalaryPayment(
            user_id=user.id, paid_on=pay_row.paid_on,
            period_year=pay_row.period_year, period_month=pay_row.period_month,
            amount=pay_row.amount, note=pay_row.note,
            recorded_by_id=user.id,
        ))
        existing_payments.add(pay_key)
        payments_inserted += 1

    return RestoreResult(
        days_inserted=days_inserted,
        days_skipped=days_skipped,
        advances_inserted=advances_inserted,
        advances_skipped=advances_skipped,
        payments_inserted=payments_inserted,
        payments_skipped=payments_skipped,
    )
