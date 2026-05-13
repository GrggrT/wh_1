"""Phase 6.11b: XLSX renderer for ``ReportData``.

Builds a single-sheet workbook with one row per period plus a totals row.
Money cells are written as floats with ``0.00`` number format so the
recipient can keep using Excel/Google Sheets formulas. Status is rendered
as the same emoji tag the text formatter uses, for at-a-glance parity.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.models import User
from src.services.reports.service import ReportData

_RU_MONTHS: tuple[str, ...] = (
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
)

_STATUS_TAGS: dict[str, str] = {
    "settled": "✅ закрыт",
    "pending": "⏳ ожидает",
    "partial": "🟡 частично",
    "overpaid": "🟢 переплата",
    "unpriced": "❔ без ставки",
}

_STATUS_FILLS: dict[str, str] = {
    "settled": "D1FAE5",   # emerald-100
    "pending": "FEE2E2",   # red-100
    "partial": "FEF3C7",   # amber-100
    "overpaid": "DBEAFE",  # blue-100
    "unpriced": "F3F4F6",  # gray-100
}

_MONEY_FMT = "#,##0.00"
_HOURS_FMT = "0.00"

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # gray-800
_TOTALS_FILL = PatternFill("solid", fgColor="F9FAFB")  # gray-50
_STRIPE_FILL = PatternFill("solid", fgColor="FAFBFC")
_BORDER_COLOR = "E5E7EB"  # gray-200
_THIN = Side(style="thin", color=_BORDER_COLOR)
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _money(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def build_report_xlsx(data: ReportData, user: User) -> BytesIO:
    """Render ``data`` into an in-memory ``.xlsx`` buffer."""
    cur = user.currency
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Отчёт"
    ws.sheet_view.showGridLines = False

    title_font = Font(bold=True, size=14, color="1F2937")
    subtitle_font = Font(size=10, color="6B7280", italic=True)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    body_font = Font(size=11, color="1F2937")
    totals_font = Font(bold=True, size=11, color="1F2937")
    centre = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", indent=1)
    right = Alignment(horizontal="right", vertical="center", indent=1)

    # Title row
    title_cell = ws.cell(
        row=1, column=1,
        value=f"Отчёт за последние {data.months} мес.",
    )
    title_cell.font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Subtitle row 2 — user + currency context
    sub_cell = ws.cell(
        row=2, column=1,
        value=f"Сотрудник: {user.name or '—'}  ·  Валюта: {cur}",
    )
    sub_cell.font = subtitle_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    # Row heights for a roomier feel
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 22

    # Header row
    headers = [
        "Период",
        "Часы",
        f"Начислено ({cur})",
        f"Получено ({cur})",
        f"Остаток ({cur})",
        "Статус",
    ]
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = header_font
        cell.fill = _HEADER_FILL
        cell.alignment = centre
        cell.border = _CELL_BORDER

    # Per-period rows
    row_idx = 4
    for i, led in enumerate(data.ledgers):
        stripe = _STRIPE_FILL if i % 2 == 1 else None
        period = f"{_RU_MONTHS[led.month - 1]} {led.year}"
        period_cell = ws.cell(row=row_idx, column=1, value=period)
        period_cell.font = body_font
        period_cell.alignment = left

        hours_cell = ws.cell(row=row_idx, column=2, value=float(led.hours))
        hours_cell.number_format = _HOURS_FMT
        hours_cell.font = body_font
        hours_cell.alignment = right

        earned = _money(led.earnings)
        if earned is not None:
            c = ws.cell(row=row_idx, column=3, value=earned)
            c.number_format = _MONEY_FMT
        else:
            c = ws.cell(row=row_idx, column=3, value="—")
            c.alignment = centre
        c.font = body_font
        if c.alignment.horizontal != "center":
            c.alignment = right

        received = _money(led.received_total)
        if received is not None:
            c = ws.cell(row=row_idx, column=4, value=received)
            c.number_format = _MONEY_FMT
            c.alignment = right
            c.font = body_font

        remaining = _money(led.remaining)
        if remaining is not None:
            c = ws.cell(row=row_idx, column=5, value=remaining)
            c.number_format = _MONEY_FMT
            c.alignment = right
        else:
            c = ws.cell(row=row_idx, column=5, value="—")
            c.alignment = centre
        c.font = body_font

        status_cell = ws.cell(
            row=row_idx, column=6,
            value=_STATUS_TAGS.get(led.status, led.status),
        )
        status_cell.font = body_font
        status_cell.alignment = centre
        fill_color = _STATUS_FILLS.get(led.status)
        if fill_color is not None:
            status_cell.fill = PatternFill("solid", fgColor=fill_color)

        # Apply alternating stripes + borders across the row
        for col in range(1, 7):
            row_cell = ws.cell(row=row_idx, column=col)
            row_cell.border = _CELL_BORDER
            if (
                stripe is not None
                and col != 6
                and row_cell.fill.fgColor.rgb in (None, "00000000")
            ):
                # Don't override the status fill on stripe rows; default cells
                # have rgb "00000000" before any fill is assigned.
                row_cell.fill = stripe
        row_idx += 1

    # Totals row
    total_row = row_idx + 1
    label_cell = ws.cell(row=total_row, column=1, value="Итого")
    label_cell.font = totals_font
    label_cell.alignment = left

    total_hours = ws.cell(
        row=total_row, column=2, value=float(data.total_hours),
    )
    total_hours.number_format = _HOURS_FMT
    total_hours.font = totals_font
    total_hours.alignment = right

    if data.total_earnings is not None:
        c = ws.cell(
            row=total_row, column=3, value=float(data.total_earnings),
        )
        c.number_format = _MONEY_FMT
        c.alignment = right
    else:
        c = ws.cell(row=total_row, column=3, value="—")
        c.alignment = centre
    c.font = totals_font

    c = ws.cell(row=total_row, column=4, value=float(data.total_received))
    c.number_format = _MONEY_FMT
    c.font = totals_font
    c.alignment = right

    c = ws.cell(row=total_row, column=5, value=float(data.total_owed))
    c.number_format = _MONEY_FMT
    c.font = totals_font
    c.alignment = right

    status_total = ws.cell(row=total_row, column=6, value=f"долг ({cur})")
    status_total.font = totals_font
    status_total.alignment = centre

    for col in range(1, 7):
        tot_cell = ws.cell(row=total_row, column=col)
        tot_cell.fill = _TOTALS_FILL
        tot_cell.border = _CELL_BORDER

    if data.total_overpaid > 0:
        over_row = total_row + 1
        ov_label = ws.cell(row=over_row, column=1, value="Переплата")
        ov_label.font = totals_font
        ov_label.alignment = left
        c = ws.cell(
            row=over_row, column=5, value=float(data.total_overpaid),
        )
        c.number_format = _MONEY_FMT
        c.font = totals_font
        c.alignment = right
        ov_tag = ws.cell(
            row=over_row, column=6, value=f"переплата ({cur})",
        )
        ov_tag.font = totals_font
        ov_tag.alignment = centre
        for col in range(1, 7):
            ov_cell = ws.cell(row=over_row, column=col)
            ov_cell.fill = _TOTALS_FILL
            ov_cell.border = _CELL_BORDER

    # Sensible column widths
    widths = (20, 12, 18, 18, 18, 22)
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze the title + subtitle + header so the data scrolls under them
    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def xlsx_filename(months: int) -> str:
    return f"report_{months}m.xlsx"
