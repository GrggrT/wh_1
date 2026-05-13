"""Phase 7.1: full-data XLSX backup.

Dumps every row owned by a single user across the accounting tables
(``day_entries``, ``advances``, ``salary_payments``) plus a one-row profile
sheet, into a single ``.xlsx`` workbook. Pure render — the caller fetches
rows and passes them in.

Layout constraints (the restore parser reads by column index starting at
``min_row=2``): header on row 1, data on row 2+, column order frozen.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.models import Advance, DayEntry, SalaryPayment, User

_MONEY_FMT = "#,##0.00"
_DATE_FMT = "yyyy-mm-dd"

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # gray-800
_STRIPE_FILL = PatternFill("solid", fgColor="F9FAFB")  # gray-50
_PROFILE_LABEL_FILL = PatternFill("solid", fgColor="F3F4F6")  # gray-100
_BORDER_COLOR = "E5E7EB"  # gray-200
_THIN = Side(style="thin", color=_BORDER_COLOR)
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_BODY_FONT = Font(size=11, color="1F2937")
_LABEL_FONT = Font(bold=True, size=11, color="1F2937")

_ALIGN_CENTRE = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", indent=1)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", indent=1)


def _money(value: Decimal | None) -> float | None:
    return None if value is None else float(value)


def _write_header(ws: Worksheet, headers: tuple[str, ...]) -> None:
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTRE
        cell.border = _CELL_BORDER
    ws.row_dimensions[1].height = 22


def _set_widths(ws: Worksheet, widths: tuple[int, ...]) -> None:
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _stripe_body(
    ws: Worksheet, *, n_rows: int, n_cols: int,
    align_by_col: dict[int, Alignment] | None = None,
) -> None:
    """Apply alternating fill + thin borders to body rows (row 2..n_rows+1)."""
    align_by_col = align_by_col or {}
    for i in range(n_rows):
        row_idx = i + 2
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = _CELL_BORDER
            cell.font = _BODY_FONT
            cell.alignment = align_by_col.get(col, _ALIGN_LEFT)
            if i % 2 == 1:
                cell.fill = _STRIPE_FILL


def _fill_profile(ws: Worksheet, user: User, today: date) -> None:
    ws.sheet_view.showGridLines = False
    rows: list[tuple[str, int | str | float | None]] = [
        ("user_id", user.id),
        ("tg_id", user.tg_id),
        ("name", user.name),
        ("currency", user.currency),
        ("hourly_rate", _money(user.hourly_rate)),
        ("role", user.role),
        ("backup_date", today.isoformat()),
    ]
    for row_idx, (label, value) in enumerate(rows, 1):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = _LABEL_FONT
        label_cell.fill = _PROFILE_LABEL_FILL
        label_cell.alignment = _ALIGN_LEFT
        label_cell.border = _CELL_BORDER

        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = _BODY_FONT
        value_cell.alignment = _ALIGN_LEFT
        value_cell.border = _CELL_BORDER
        if label == "hourly_rate" and isinstance(value, float):
            value_cell.number_format = _MONEY_FMT
            value_cell.alignment = _ALIGN_RIGHT
        ws.row_dimensions[row_idx].height = 20
    _set_widths(ws, (22, 28))


def _fill_day_entries(ws: Worksheet, entries: list[DayEntry]) -> None:
    ws.sheet_view.showGridLines = False
    _write_header(ws, ("id", "day", "hours", "site_id", "note", "created_at"))
    align_by_col = {
        1: _ALIGN_CENTRE,
        2: _ALIGN_CENTRE,
        3: _ALIGN_RIGHT,
        4: _ALIGN_CENTRE,
        5: _ALIGN_LEFT,
        6: _ALIGN_CENTRE,
    }
    for row_idx, e in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=e.id)
        c = ws.cell(row=row_idx, column=2, value=e.day)
        c.number_format = _DATE_FMT
        c = ws.cell(row=row_idx, column=3, value=float(e.hours))
        c.number_format = _MONEY_FMT
        ws.cell(row=row_idx, column=4, value=e.site_id)
        ws.cell(row=row_idx, column=5, value=e.note)
        ws.cell(
            row=row_idx, column=6,
            value=e.created_at.isoformat() if e.created_at else None,
        )
    _stripe_body(
        ws, n_rows=len(entries), n_cols=6, align_by_col=align_by_col,
    )
    _set_widths(ws, (8, 14, 12, 12, 42, 26))
    ws.freeze_panes = "A2"


def _fill_advances(ws: Worksheet, advances: list[Advance]) -> None:
    ws.sheet_view.showGridLines = False
    _write_header(
        ws,
        ("id", "day", "period_year", "period_month", "amount", "note", "created_at"),
    )
    align_by_col = {
        1: _ALIGN_CENTRE,
        2: _ALIGN_CENTRE,
        3: _ALIGN_CENTRE,
        4: _ALIGN_CENTRE,
        5: _ALIGN_RIGHT,
        6: _ALIGN_LEFT,
        7: _ALIGN_CENTRE,
    }
    for row_idx, a in enumerate(advances, 2):
        ws.cell(row=row_idx, column=1, value=a.id)
        c = ws.cell(row=row_idx, column=2, value=a.day)
        c.number_format = _DATE_FMT
        ws.cell(row=row_idx, column=3, value=a.period_year)
        ws.cell(row=row_idx, column=4, value=a.period_month)
        c = ws.cell(row=row_idx, column=5, value=float(a.amount))
        c.number_format = _MONEY_FMT
        ws.cell(row=row_idx, column=6, value=a.note)
        ws.cell(
            row=row_idx, column=7,
            value=a.created_at.isoformat() if a.created_at else None,
        )
    _stripe_body(
        ws, n_rows=len(advances), n_cols=7, align_by_col=align_by_col,
    )
    _set_widths(ws, (8, 14, 13, 14, 14, 42, 26))
    ws.freeze_panes = "A2"


def _fill_payments(ws: Worksheet, payments: list[SalaryPayment]) -> None:
    ws.sheet_view.showGridLines = False
    _write_header(
        ws,
        (
            "id", "paid_on", "period_year", "period_month",
            "amount", "note", "created_at",
        ),
    )
    align_by_col = {
        1: _ALIGN_CENTRE,
        2: _ALIGN_CENTRE,
        3: _ALIGN_CENTRE,
        4: _ALIGN_CENTRE,
        5: _ALIGN_RIGHT,
        6: _ALIGN_LEFT,
        7: _ALIGN_CENTRE,
    }
    for row_idx, p in enumerate(payments, 2):
        ws.cell(row=row_idx, column=1, value=p.id)
        c = ws.cell(row=row_idx, column=2, value=p.paid_on)
        c.number_format = _DATE_FMT
        ws.cell(row=row_idx, column=3, value=p.period_year)
        ws.cell(row=row_idx, column=4, value=p.period_month)
        c = ws.cell(row=row_idx, column=5, value=float(p.amount))
        c.number_format = _MONEY_FMT
        ws.cell(row=row_idx, column=6, value=p.note)
        ws.cell(
            row=row_idx, column=7,
            value=p.created_at.isoformat() if p.created_at else None,
        )
    _stripe_body(
        ws, n_rows=len(payments), n_cols=7, align_by_col=align_by_col,
    )
    _set_widths(ws, (8, 14, 13, 14, 14, 42, 26))
    ws.freeze_panes = "A2"


def build_backup_xlsx(
    user: User,
    day_entries: list[DayEntry],
    advances: list[Advance],
    payments: list[SalaryPayment],
    today: date,
) -> BytesIO:
    """Build a 4-sheet workbook: profile + day_entries + advances + payments."""
    wb = Workbook()
    # First sheet is created by openpyxl; rename and use it as Profile.
    ws_profile = wb.active
    assert ws_profile is not None
    ws_profile.title = "Профиль"
    _fill_profile(ws_profile, user, today)

    ws_days = wb.create_sheet("Дни")
    _fill_day_entries(ws_days, day_entries)

    ws_adv = wb.create_sheet("Авансы")
    _fill_advances(ws_adv, advances)

    ws_pay = wb.create_sheet("Выплаты")
    _fill_payments(ws_pay, payments)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def backup_filename(user: User, today: date | datetime) -> str:
    d = today.date() if isinstance(today, datetime) else today
    return f"wh1_backup_{user.id}_{d.isoformat()}.xlsx"
