"""XLSX timesheet exporter."""

from datetime import date, time
from decimal import Decimal
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.models import Break, Crew, Shift, Site, User
from src.services.breaks import total_break_hours
from src.services.reports import compute_hours, split_shift_at_midnight

_MONEY_FMT = "#,##0.00"
_HOURS_FMT = "0.00"

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # gray-800
_STRIPE_FILL = PatternFill("solid", fgColor="F9FAFB")  # gray-50
_TOTALS_FILL = PatternFill("solid", fgColor="F3F4F6")  # gray-100
_LABEL_FILL = PatternFill("solid", fgColor="E5E7EB")  # gray-200
_BORDER_COLOR = "E5E7EB"
_THIN = Side(style="thin", color=_BORDER_COLOR)
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_BODY_FONT = Font(size=11, color="1F2937")
_LABEL_FONT = Font(bold=True, size=11, color="1F2937")
_TITLE_FONT = Font(bold=True, size=12, color="1F2937")

_ALIGN_CENTRE = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", indent=1)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", indent=1)


class ShiftRow:
    """Represents one row in the exported timesheet (after midnight splits)."""

    def __init__(
        self,
        shift_date: date,
        site_name: str,
        start_time: time,
        end_time: time,
        hours: Decimal,
        rate: Decimal | None,
        note: str | None,
    ) -> None:
        self.shift_date = shift_date
        self.site_name = site_name
        self.start_time = start_time
        self.end_time = end_time
        self.hours = hours
        self.rate = rate
        self.amount = hours * rate if rate else None
        self.note = note


def build_shift_rows(
    shifts: list[Shift],
    sites: dict[int, Site],
    user: User,
    tz: ZoneInfo,
    breaks_by_shift: dict[int, list[Break]] | None = None,
) -> list[ShiftRow]:
    """Build export rows, splitting shifts at midnight; subtract breaks per segment."""
    rows: list[ShiftRow] = []

    for shift in shifts:
        if shift.end_at is None:
            continue

        site = sites.get(shift.site_id) if shift.site_id else None
        site_name = site.name if site else ""

        rate: Decimal | None = None
        if site and site.hourly_rate is not None:
            rate = site.hourly_rate
        elif user.hourly_rate is not None:
            rate = user.hourly_rate

        shift_breaks = (
            (breaks_by_shift or {}).get(shift.id, []) if breaks_by_shift else []
        )

        segments = split_shift_at_midnight(shift.start_at, shift.end_at, tz)
        for seg_start, seg_end in segments:
            gross = compute_hours(seg_start, seg_end)
            break_h = (
                total_break_hours(shift_breaks, seg_start, seg_end)
                if shift_breaks else Decimal(0)
            )
            net = gross - break_h
            if net < 0:
                net = Decimal(0)
            rows.append(
                ShiftRow(
                    shift_date=seg_start.date(),
                    site_name=site_name,
                    start_time=seg_start.timetz(),
                    end_time=seg_end.timetz(),
                    hours=net.quantize(Decimal("0.01")),
                    rate=rate,
                    note=shift.note,
                )
            )

    return rows


def _write_shifts_header(ws: Worksheet, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTRE
        cell.border = _CELL_BORDER
    ws.row_dimensions[1].height = 22


def _write_shift_row(ws: Worksheet, row_idx: int, row: ShiftRow) -> None:
    stripe = row_idx % 2 == 1  # row 3, 5, 7… get stripe
    align_by_col = {
        1: _ALIGN_CENTRE, 2: _ALIGN_LEFT,
        3: _ALIGN_CENTRE, 4: _ALIGN_CENTRE,
        5: _ALIGN_RIGHT, 6: _ALIGN_RIGHT,
        7: _ALIGN_RIGHT, 8: _ALIGN_LEFT,
    }

    ws.cell(row=row_idx, column=1, value=row.shift_date.isoformat())
    ws.cell(row=row_idx, column=2, value=row.site_name)
    ws.cell(row=row_idx, column=3, value=row.start_time.strftime("%H:%M"))
    ws.cell(row=row_idx, column=4, value=row.end_time.strftime("%H:%M"))

    hours_cell = ws.cell(row=row_idx, column=5, value=float(row.hours))
    hours_cell.number_format = _HOURS_FMT

    if row.rate is not None:
        rate_cell = ws.cell(row=row_idx, column=6, value=float(row.rate))
        rate_cell.number_format = _MONEY_FMT

    if row.amount is not None:
        amount_cell = ws.cell(row=row_idx, column=7, value=float(row.amount))
        amount_cell.number_format = _MONEY_FMT

    ws.cell(row=row_idx, column=8, value=row.note or "")

    for col in range(1, 9):
        c = ws.cell(row=row_idx, column=col)
        c.font = _BODY_FONT
        c.alignment = align_by_col.get(col, _ALIGN_LEFT)
        c.border = _CELL_BORDER
        if stripe:
            c.fill = _STRIPE_FILL


def _set_widths(ws: Worksheet, widths: tuple[int, ...]) -> None:
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def export_xlsx(
    shifts: list[Shift],
    sites: dict[int, Site],
    user: User,
    tz: ZoneInfo,
    period: str,
    breaks_by_shift: dict[int, list[Break]] | None = None,
) -> BytesIO:
    """Generate XLSX workbook and return as BytesIO."""
    rows = build_shift_rows(shifts, sites, user, tz, breaks_by_shift)
    wb = Workbook()

    # Sheet 1 - Shifts
    ws = wb.active
    assert ws is not None
    ws.title = "Shifts"
    ws.sheet_view.showGridLines = False

    headers = [
        "Date", "Site", "Start", "End",
        "Hours", "Rate (zl)", "Amount (zl)", "Note",
    ]
    _write_shifts_header(ws, headers)

    for row_idx, row in enumerate(rows, 2):
        _write_shift_row(ws, row_idx, row)

    _set_widths(ws, (12, 22, 9, 9, 10, 12, 14, 30))
    ws.freeze_panes = "A2"

    # Sheet 2 - Summary
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_view.showGridLines = False

    def _section_title(row_idx: int, text: str) -> None:
        cell = ws_sum.cell(row=row_idx, column=1, value=text)
        cell.font = _TITLE_FONT
        cell.fill = _LABEL_FILL
        cell.alignment = _ALIGN_LEFT
        cell.border = _CELL_BORDER
        ws_sum.cell(row=row_idx, column=2).fill = _LABEL_FILL
        ws_sum.cell(row=row_idx, column=2).border = _CELL_BORDER

    def _value_row(row_idx: int, label: str, value: float, fmt: str) -> None:
        lbl = ws_sum.cell(row=row_idx, column=1, value=label)
        lbl.font = _BODY_FONT
        lbl.alignment = _ALIGN_LEFT
        lbl.border = _CELL_BORDER
        val = ws_sum.cell(row=row_idx, column=2, value=value)
        val.font = _BODY_FONT
        val.alignment = _ALIGN_RIGHT
        val.border = _CELL_BORDER
        val.number_format = fmt

    # Block 1: Hours per site
    _section_title(1, "Hours per site")
    site_hours: dict[str, Decimal] = {}
    for row in rows:
        site_hours[row.site_name] = site_hours.get(row.site_name, Decimal(0)) + row.hours

    sum_row = 2
    for site_name, hours in sorted(site_hours.items()):
        _value_row(sum_row, site_name, float(hours), _HOURS_FMT)
        sum_row += 1

    # Block 2: Hours per work_type
    sum_row += 1
    _section_title(sum_row, "Hours per work type")
    sum_row += 1
    type_hours: dict[str, Decimal] = {}
    for shift in shifts:
        if shift.end_at is None:
            continue
        wt = shift.work_type or "(not specified)"
        h = compute_hours(shift.start_at, shift.end_at)
        type_hours[wt] = type_hours.get(wt, Decimal(0)) + h

    for wt, hours in sorted(type_hours.items()):
        _value_row(
            sum_row, wt, float(hours.quantize(Decimal("0.01"))), _HOURS_FMT,
        )
        sum_row += 1

    # Block 3: Grand totals
    sum_row += 1
    _section_title(sum_row, "Grand total")
    sum_row += 1
    grand_hours = sum(r.hours for r in rows)
    grand_amount = sum(r.amount for r in rows if r.amount is not None)

    _value_row(sum_row, "Total hours", float(grand_hours), _HOURS_FMT)
    sum_row += 1

    _value_row(sum_row, "Total amount (zl)", float(grand_amount), _MONEY_FMT)
    # Highlight the grand-total numbers
    for r in (sum_row - 1, sum_row):
        for col in (1, 2):
            ws_sum.cell(row=r, column=col).fill = _TOTALS_FILL
            ws_sum.cell(row=r, column=col).font = _LABEL_FONT

    _set_widths(ws_sum, (32, 18))

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_crew_xlsx(
    crew: Crew,
    members: list[User],
    shifts_by_user: dict[int, list[Shift]],
    sites: dict[int, Site],
    tz: ZoneInfo,
    period: str,
    breaks_by_shift: dict[int, list[Break]] | None = None,
) -> BytesIO:
    """Generate a crew-wide workbook: one sheet per member + summary."""
    wb = Workbook()

    # Summary sheet first — keep layout: row 1 = crew/period, row 3 = headers,
    # row 4+ = member rows, last = "Grand total".
    ws_sum = wb.active
    assert ws_sum is not None
    ws_sum.title = "Crew summary"
    ws_sum.sheet_view.showGridLines = False

    crew_cell = ws_sum.cell(row=1, column=1, value=f"Crew: {crew.name}")
    crew_cell.font = _TITLE_FONT
    crew_cell.alignment = _ALIGN_LEFT
    period_cell = ws_sum.cell(row=1, column=2, value=f"Period: {period}")
    period_cell.font = Font(size=11, italic=True, color="6B7280")
    period_cell.alignment = _ALIGN_LEFT
    ws_sum.row_dimensions[1].height = 22

    header_labels = ("Member", "Hours", "Amount (zl)", "Shifts")
    for col, label in enumerate(header_labels, 1):
        cell = ws_sum.cell(row=3, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTRE
        cell.border = _CELL_BORDER
    ws_sum.row_dimensions[3].height = 22

    summary_row = 4
    grand_hours = Decimal(0)
    grand_amount = Decimal(0)
    align_by_col = {
        1: _ALIGN_LEFT, 2: _ALIGN_RIGHT, 3: _ALIGN_RIGHT, 4: _ALIGN_CENTRE,
    }
    for i, member in enumerate(members):
        member_shifts = shifts_by_user.get(member.id, [])
        rows = build_shift_rows(member_shifts, sites, member, tz, breaks_by_shift)
        member_hours = sum((r.hours for r in rows), Decimal(0))
        member_amount = sum(
            (r.amount for r in rows if r.amount is not None), Decimal(0),
        )
        ws_sum.cell(row=summary_row, column=1, value=member.name)
        c = ws_sum.cell(row=summary_row, column=2, value=float(member_hours))
        c.number_format = _HOURS_FMT
        c = ws_sum.cell(row=summary_row, column=3, value=float(member_amount))
        c.number_format = _MONEY_FMT
        ws_sum.cell(
            row=summary_row, column=4,
            value=len([s for s in member_shifts if s.end_at is not None]),
        )
        stripe = i % 2 == 1
        for col in range(1, 5):
            mem_cell = ws_sum.cell(row=summary_row, column=col)
            mem_cell.font = _BODY_FONT
            mem_cell.alignment = align_by_col.get(col, _ALIGN_LEFT)
            mem_cell.border = _CELL_BORDER
            if stripe:
                mem_cell.fill = _STRIPE_FILL
        summary_row += 1
        grand_hours += member_hours
        grand_amount += member_amount

        # Per-member sheet
        sheet_name = (member.name or f"user_{member.id}")[:31]
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        headers = [
            "Date", "Site", "Start", "End",
            "Hours", "Rate (zl)", "Amount (zl)", "Note",
        ]
        _write_shifts_header(ws, headers)
        for row_idx, row in enumerate(rows, 2):
            _write_shift_row(ws, row_idx, row)
        _set_widths(ws, (12, 22, 9, 9, 10, 12, 14, 30))
        ws.freeze_panes = "A2"

    summary_row += 1
    total_label = ws_sum.cell(row=summary_row, column=1, value="Grand total")
    total_label.font = _LABEL_FONT
    total_label.alignment = _ALIGN_LEFT
    c = ws_sum.cell(row=summary_row, column=2, value=float(grand_hours))
    c.number_format = _HOURS_FMT
    c.font = _LABEL_FONT
    c.alignment = _ALIGN_RIGHT
    c = ws_sum.cell(row=summary_row, column=3, value=float(grand_amount))
    c.number_format = _MONEY_FMT
    c.font = _LABEL_FONT
    c.alignment = _ALIGN_RIGHT
    ws_sum.cell(row=summary_row, column=4).font = _LABEL_FONT
    for col in range(1, 5):
        ws_sum.cell(row=summary_row, column=col).fill = _TOTALS_FILL
        ws_sum.cell(row=summary_row, column=col).border = _CELL_BORDER

    _set_widths(ws_sum, (22, 12, 16, 10))
    ws_sum.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
